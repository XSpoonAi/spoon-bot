#!/usr/bin/env python3
"""Document export skill script.

Reads JSON from stdin, exports content to the requested format,
writes JSON result to stdout.

Input:  {"content": "...", "format": "markdown|pdf|excel|mermaid", "title": "...", "output_path": "..."}
Output: {"success": true, "file_path": "...", "message": "..."}
"""

from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any


def default_output_path(output_format: str, exports_dir: Path) -> Path:
    """Generate a default output path based on format."""
    exports_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    ext_map = {"excel": "xlsx", "markdown": "md", "mermaid": "md", "pdf": "pdf"}
    return exports_dir / f"export_{stamp}.{ext_map.get(output_format, 'txt')}"


def export_markdown(content: str, output: Path) -> str:
    """Export content as markdown."""
    markdown_text = content
    try:
        parsed = json.loads(content)
        markdown_text = f"```json\n{json.dumps(parsed, indent=2, ensure_ascii=False)}\n```"
    except Exception:
        pass
    output.write_text(markdown_text, encoding="utf-8")
    return f"Exported markdown: {output}"


def export_mermaid(content: str, output: Path, title: str | None = None) -> str:
    """Export content as a mermaid mindmap."""
    doc_title = title or "Document"
    lines = [line.strip("- ") for line in content.splitlines() if line.strip()]
    mermaid = ["mindmap", f"  root(({doc_title}))"]
    for line in lines[:30]:
        mermaid.append(f"    {line}")
    output.write_text("```mermaid\n" + "\n".join(mermaid) + "\n```\n", encoding="utf-8")
    return f"Exported mermaid: {output}"


def export_excel(
    content: str, output: Path, title: str | None = None, sheets: str | None = None
) -> str:
    """Export content as an Excel spreadsheet."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "Error: openpyxl is required for Excel export. Install with: pip install openpyxl"

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    sheet_defs: list[dict[str, Any]] = []
    if sheets:
        try:
            parsed_sheets = json.loads(sheets)
            if isinstance(parsed_sheets, list):
                for item in parsed_sheets:
                    if isinstance(item, dict):
                        sheet_defs.append(item)
        except Exception:
            return "Error: Invalid sheets JSON"

    if not sheet_defs:
        sheet_defs.append({"name": title or "Sheet1", "data": content})

    for sheet_def in sheet_defs:
        ws = workbook.create_sheet(title=str(sheet_def.get("name", "Sheet"))[:31])
        data = sheet_def.get("data", "")
        rows: list[list[str]] = []
        for line in str(data).splitlines():
            if "|" in line:
                cells = [cell.strip() for cell in line.split("|") if cell.strip()]
                if cells:
                    rows.append(cells)
        if not rows:
            rows = [[str(data)]]

        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)

        for c_idx in range(1, max(len(r) for r in rows) + 1):
            col_values = [str(r[c_idx - 1]) for r in rows if len(r) >= c_idx]
            width = min(60, max(10, max(len(v) for v in col_values) + 2))
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    workbook.save(output)
    return f"Exported excel: {output}"


def export_pdf(content: str, output: Path, title: str | None = None) -> str:
    """Export content as a PDF document."""
    try:
        import fitz  # type: ignore
    except ImportError:
        return "Error: pymupdf is required for PDF export. Install with: pip install pymupdf"

    pdf_doc = fitz.open()
    page = pdf_doc.new_page()
    y = 72
    if title:
        page.insert_text((72, y), title, fontsize=16)
        y += 24
    for line in content.splitlines() or [content]:
        if y > 800:
            page = pdf_doc.new_page()
            y = 72
        page.insert_text((72, y), line, fontsize=11)
        y += 14
    pdf_doc.save(str(output))
    pdf_doc.close()
    return f"Exported pdf: {output}"


def main() -> None:
    """Main entry point: read JSON from stdin, export, write JSON to stdout."""
    try:
        raw = sys.stdin.read()
        params = json.loads(raw)
    except (json.JSONDecodeError, Exception) as e:
        json.dump({"success": False, "error": f"Invalid JSON input: {e}"}, sys.stdout)
        return

    content = params.get("content", "")
    out_format = params.get("format", "").lower().strip()
    title = params.get("title")
    output_path = params.get("output_path")
    sheets = params.get("sheets")

    if not content:
        json.dump({"success": False, "error": "Missing required field: content"}, sys.stdout)
        return

    if out_format not in {"excel", "markdown", "mermaid", "pdf"}:
        json.dump(
            {"success": False, "error": "Invalid format. Supported: excel, markdown, mermaid, pdf"},
            sys.stdout,
        )
        return

    exports_dir = Path("workspace/exports")
    output = Path(output_path).expanduser().resolve() if output_path else default_output_path(out_format, exports_dir)
    output.parent.mkdir(parents=True, exist_ok=True)

    try:
        if out_format == "markdown":
            message = export_markdown(content, output)
        elif out_format == "mermaid":
            message = export_mermaid(content, output, title)
        elif out_format == "excel":
            message = export_excel(content, output, title, sheets)
        elif out_format == "pdf":
            message = export_pdf(content, output, title)
        else:
            message = "Error: Unsupported format"

        if message.startswith("Error:"):
            json.dump({"success": False, "error": message}, sys.stdout)
        else:
            json.dump({"success": True, "file_path": str(output), "message": message}, sys.stdout)
    except Exception as exc:
        json.dump({"success": False, "error": f"Export failed: {exc}"}, sys.stdout)


if __name__ == "__main__":
    main()
