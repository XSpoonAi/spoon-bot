"""Document processing tools for parsing, exporting, and summarizing content."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from spoon_bot.agent.tools.base import Tool, ToolParameterSchema


class DocumentParseTool(Tool):
    """Parse PDF documents to extract text, tables, and images."""

    def __init__(self, workspace: str | Path = "./workspace") -> None:
        self.workspace = Path(workspace).resolve()
        self.images_dir = self.workspace / "extracted_images"

    @property
    def name(self) -> str:
        """Tool name used in function calls."""
        return "document_parse"

    @property
    def description(self) -> str:
        """Tool description for agent selection."""
        return "Parse PDF documents and extract text, tables, and images with page controls."

    @property
    def parameters(self) -> ToolParameterSchema:
        """JSON schema for parse parameters."""
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Path to PDF file",
                },
                "extract_text": {
                    "type": "boolean",
                    "description": "Extract plain text content",
                    "default": True,
                },
                "extract_tables": {
                    "type": "boolean",
                    "description": "Extract detected tables",
                    "default": True,
                },
                "extract_images": {
                    "type": "boolean",
                    "description": "Extract embedded images",
                    "default": False,
                },
                "page_range": {
                    "type": "string",
                    "description": "Pages like '1-10' or '1,3,5-7'",
                },
            },
            "required": ["file_path"],
        }

    def _parse_page_range(self, page_range: str | None, page_count: int) -> list[int]:
        if not page_range:
            return list(range(page_count))

        pages: set[int] = set()
        parts = [part.strip() for part in page_range.split(",") if part.strip()]
        for part in parts:
            if "-" in part:
                bounds = [p.strip() for p in part.split("-", 1)]
                if len(bounds) != 2:
                    continue
                start = int(bounds[0])
                end = int(bounds[1])
                if start > end:
                    start, end = end, start
                for page_num in range(start, end + 1):
                    if 1 <= page_num <= page_count:
                        pages.add(page_num - 1)
            else:
                page_num = int(part)
                if 1 <= page_num <= page_count:
                    pages.add(page_num - 1)

        if not pages:
            return []
        return sorted(pages)

    async def execute(
        self,
        file_path: str,
        extract_text: bool = True,
        extract_tables: bool = True,
        extract_images: bool = False,
        page_range: str | None = None,
        **_: Any,
    ) -> str:
        """Parse PDF and return structured extraction output."""
        try:
            import fitz  # type: ignore
        except ImportError:
            return "Error: pymupdf is required for PDF parsing. Install with: pip install pymupdf"

        pdf_path = Path(file_path).expanduser().resolve()
        if not pdf_path.exists():
            return f"Error: File not found: {file_path}"
        if pdf_path.suffix.lower() != ".pdf":
            return "Error: Only PDF files are supported"

        try:
            doc = fitz.open(str(pdf_path))
        except Exception as exc:
            return f"Error: Failed to open PDF: {exc}"

        try:
            if getattr(doc, "needs_pass", False):
                return "Error: PDF is password-protected"

            page_count = len(doc)
            try:
                selected_pages = self._parse_page_range(page_range, page_count)
            except Exception:
                return "Error: Invalid page_range format. Use like '1-10' or '1,3,5-7'"

            if not selected_pages:
                return "Error: No valid pages selected"

            warnings: list[str] = []
            if page_count > 100:
                warnings.append("Warning: Document has more than 100 pages")

            text_blocks: list[str] = []
            table_blocks: list[str] = []
            image_paths: list[str] = []

            if extract_images:
                self.images_dir.mkdir(parents=True, exist_ok=True)

            for page_index in selected_pages:
                page = doc[page_index]
                page_label = page_index + 1

                if extract_text:
                    text = page.get_text("text") or ""
                    if text.strip():
                        text_blocks.append(f"## Page {page_label}\n{text.strip()}")

                if extract_tables:
                    try:
                        tables = page.find_tables()
                        table_items = getattr(tables, "tables", []) if tables else []
                        for idx, table in enumerate(table_items, start=1):
                            rows = table.extract() if hasattr(table, "extract") else []
                            table_blocks.append(
                                f"## Page {page_label} Table {idx}\n{json.dumps(rows, ensure_ascii=False)}"
                            )
                    except Exception:
                        pass

                if extract_images:
                    try:
                        images = page.get_images(full=True)
                        for idx, img in enumerate(images, start=1):
                            xref = img[0]
                            image_info = doc.extract_image(xref)
                            image_bytes = image_info.get("image")
                            image_ext = image_info.get("ext", "png")
                            image_name = f"page_{page_label}_img_{idx}.{image_ext}"
                            image_path = self.images_dir / image_name
                            if image_bytes:
                                image_path.write_bytes(image_bytes)
                                image_paths.append(str(image_path))
                    except Exception:
                        pass

            result_parts: list[str] = [
                f"Parsed: {pdf_path.name}",
                f"Pages processed: {len(selected_pages)}/{page_count}",
            ]
            result_parts.extend(warnings)

            if extract_text:
                joined_text = "\n\n".join(text_blocks)
                if len(joined_text) > 50_000:
                    joined_text = joined_text[:50_000] + "\n\n[Truncated at 50KB]"
                result_parts.append("\n# Text\n" + (joined_text or "No text extracted"))

            if extract_tables:
                joined_tables = "\n\n".join(table_blocks)
                result_parts.append("\n# Tables\n" + (joined_tables or "No tables found"))

            if extract_images:
                images_text = "\n".join(image_paths) if image_paths else "No images extracted"
                result_parts.append("\n# Images\n" + images_text)

            return "\n".join(result_parts)
        finally:
            doc.close()


class DocumentExportTool(Tool):
    """Export content to markdown, mermaid, excel, or simple PDF."""

    def __init__(self, workspace: str | Path = "./workspace") -> None:
        self.workspace = Path(workspace).resolve()
        self.exports_dir = self.workspace / "exports"

    @property
    def name(self) -> str:
        """Tool name used in function calls."""
        return "document_export"

    @property
    def description(self) -> str:
        """Tool description for agent selection."""
        return "Export content to excel, markdown, mermaid mindmap, or text-based PDF."

    @property
    def parameters(self) -> ToolParameterSchema:
        """JSON schema for export parameters."""
        return {
            "type": "object",
            "properties": {
                "content": {"type": "string", "description": "Content to export"},
                "format": {
                    "type": "string",
                    "description": "Output format",
                    "enum": ["excel", "markdown", "mermaid", "pdf"],
                },
                "output_path": {
                    "type": "string",
                    "description": "Optional output path",
                },
                "title": {"type": "string", "description": "Document title"},
                "sheets": {
                    "type": "string",
                    "description": "JSON array of sheets [{name, data}]",
                },
            },
            "required": ["content", "format"],
        }

    def _default_output_path(self, output_format: str) -> Path:
        self.exports_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ext_map = {"excel": "xlsx", "markdown": "md", "mermaid": "md", "pdf": "pdf"}
        return self.exports_dir / f"export_{stamp}.{ext_map[output_format]}"

    async def execute(
        self,
        content: str,
        format: str,
        output_path: str | None = None,
        title: str | None = None,
        sheets: str | None = None,
        **_: Any,
    ) -> str:
        """Export content into the requested format."""
        out_format = format.lower().strip()
        if out_format not in {"excel", "markdown", "mermaid", "pdf"}:
            return "Error: Invalid format. Supported: excel, markdown, mermaid, pdf"

        output = Path(output_path).expanduser().resolve() if output_path else self._default_output_path(out_format)
        output.parent.mkdir(parents=True, exist_ok=True)

        try:
            if out_format == "markdown":
                markdown_text = content
                try:
                    parsed = json.loads(content)
                    markdown_text = f"```json\n{json.dumps(parsed, indent=2, ensure_ascii=False)}\n```"
                except Exception:
                    pass
                output.write_text(markdown_text, encoding="utf-8")
                return f"Exported markdown: {output}"

            if out_format == "mermaid":
                doc_title = title or "Document"
                lines = [line.strip("- ") for line in content.splitlines() if line.strip()]
                mermaid = ["mindmap", f"  root(({doc_title}))"]
                for line in lines[:30]:
                    mermaid.append(f"    {line}")
                output.write_text("```mermaid\n" + "\n".join(mermaid) + "\n```\n", encoding="utf-8")
                return f"Exported mermaid: {output}"

            if out_format == "excel":
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font
                    from openpyxl.utils import get_column_letter
                except ImportError:
                    return "Error: openpyxl is required for Excel export. Install with: pip install openpyxl"

                workbook = Workbook()
                default_sheet = workbook.active
                workbook.remove(default_sheet)

                sheet_defs: list[dict[str, Any]] = []
                if sheets:
                    try:
                        parsed_sheets = json.loads(sheets)
                        if isinstance(parsed_sheets, list):
                            for item in parsed_sheets:
                                if isinstance(item, dict):
                                    sheet_defs.append(item)
                    except Exception:
                        return "Error: Invalid sheets JSON"

                if not sheet_defs:
                    sheet_defs.append({"name": title or "Sheet1", "data": content})

                for sheet_def in sheet_defs:
                    ws = workbook.create_sheet(title=str(sheet_def.get("name", "Sheet"))[:31])
                    data = sheet_def.get("data", "")
                    rows: list[list[str]] = []
                    for line in str(data).splitlines():
                        if "|" in line:
                            cells = [cell.strip() for cell in line.split("|") if cell.strip()]
                            if cells:
                                rows.append(cells)
                    if not rows:
                        rows = [[str(data)]]

                    for r_idx, row in enumerate(rows, start=1):
                        for c_idx, value in enumerate(row, start=1):
                            cell = ws.cell(row=r_idx, column=c_idx, value=value)
                            if r_idx == 1:
                                cell.font = Font(bold=True)

                    for c_idx in range(1, max(len(r) for r in rows) + 1):
                        col_values = [str(r[c_idx - 1]) for r in rows if len(r) >= c_idx]
                        width = min(60, max(10, max(len(v) for v in col_values) + 2))
                        ws.column_dimensions[get_column_letter(c_idx)].width = width

                workbook.save(output)
                return f"Exported excel: {output}"

            if out_format == "pdf":
                try:
                    import fitz  # type: ignore
                except ImportError:
                    return "Error: pymupdf is required for PDF parsing. Install with: pip install pymupdf"

                pdf_doc = fitz.open()
                page = pdf_doc.new_page()
                y = 72
                if title:
                    page.insert_text((72, y), title, fontsize=16)
                    y += 24
                for line in content.splitlines() or [content]:
                    if y > 800:
                        page = pdf_doc.new_page()
                        y = 72
                    page.insert_text((72, y), line, fontsize=11)
                    y += 14
                pdf_doc.save(str(output))
                pdf_doc.close()
                return f"Exported pdf: {output}"

            return "Error: Unsupported export format"
        except Exception as exc:
            return f"Error: Export failed: {exc}"


class DocumentSummarizeTool(Tool):
    """Provide structured summarization templates for document analysis."""

    @property
    def name(self) -> str:
        """Tool name used in function calls."""
        return "document_summarize"

    @property
    def description(self) -> str:
        """Tool description for agent selection."""
        return "Return structured markdown templates for summarizing document text."

    @property
    def parameters(self) -> ToolParameterSchema:
        """JSON schema for summarize parameters."""
        return {
            "type": "object",
            "properties": {
                "text": {"type": "string", "description": "Document text to summarize"},
                "template": {
                    "type": "string",
                    "description": "Template type",
                    "enum": ["tokenomics", "roadmap", "whitepaper", "general"],
                    "default": "general",
                },
                "output_format": {
                    "type": "string",
                    "description": "Output format",
                    "default": "markdown",
                },
            },
            "required": ["text"],
        }

    async def execute(
        self,
        text: str,
        template: str = "general",
        output_format: str = "markdown",
        **_: Any,
    ) -> str:
        """Return a template the agent can fill with extracted content."""
        if not text.strip():
            return "Error: text cannot be empty"
        if output_format.lower() != "markdown":
            return "Error: Only markdown output_format is currently supported"

        template_name = template.lower().strip()
        templates: dict[str, str] = {
            "tokenomics": (
                "# Tokenomics Summary\n\n"
                "## Token Name\n- \n\n"
                "## Total Supply\n- \n\n"
                "## Distribution\n- Team: \n- Community: \n- Treasury: \n\n"
                "## Vesting Schedule\n- \n\n"
                "## Utility\n- \n"
            ),
            "roadmap": (
                "# Roadmap Summary\n\n"
                "## Phases\n- \n\n"
                "## Milestones\n- \n\n"
                "## Dates\n- \n\n"
                "## Deliverables\n- \n"
            ),
            "whitepaper": (
                "# Whitepaper Summary\n\n"
                "## Abstract\n- \n\n"
                "## Problem\n- \n\n"
                "## Solution\n- \n\n"
                "## Technology\n- \n\n"
                "## Team\n- \n\n"
                "## Tokenomics\n- \n\n"
                "## Roadmap\n- \n"
            ),
            "general": (
                "# Document Summary\n\n"
                "## Title\n- \n\n"
                "## Summary\n- \n\n"
                "## Key Points\n- \n\n"
                "## Conclusions\n- \n"
            ),
        }

        if template_name not in templates:
            return "Error: Invalid template. Supported: tokenomics, roadmap, whitepaper, general"

        preview = text.strip().replace("\n", " ")[:300]
        return templates[template_name] + f"\n---\nSource preview: {preview}"
